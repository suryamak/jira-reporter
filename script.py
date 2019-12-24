import openpyxl as xl

wb = xl.load_workbook(filename='Weekly Report Ending 12-20-2019.xlsx')
ws = wb["Jira 2019-12-20T17_47_01+0000"]
title = ws.title
headers = tuple(ws.rows)[0]
assignee_col = -1
priority_col = -1
for cell in headers:
    if (cell.value.lower() == 'assignee'):
        assignee_col = cell.column
    elif (cell.value.lower() == 'priority'):
        priority_col = cell.column

assignees = {}
priorities = []

for row in ws.iter_rows(min_row=2):
    assignee = row[assignee_col-1].value
    priority = row[priority_col-1].value
    if priority not in priorities:
        priorities.append(priority)
    if (assignee in assignees.keys()):
        if (priority in assignees[assignee].keys()):
            assignees[assignee][priority] = assignees[assignee][priority] + 1
        else:
            assignees[assignee][priority] = 1
    else:
        assignees[assignee] = {}
        assignees[assignee][priority] = 1
    

ws_out = wb.create_sheet(title=("RESULT_"+title)[0:31])

ws_out.cell(row=1, column=1, value="Assignees")
last_col = -1
for i in range(0, len(priorities)):
    ws_out.cell(row=1, column=i+2, value=priorities[i])
    last_col = i

last_col += 3

ws_out.cell(row=1, column=last_col, value="Total")

last_row = 2
for key in assignees:
    total = 0
    ws_out.cell(row=last_row, column=1, value=key)
    for level in assignees[key]:
        ws_out.cell(row=last_row, column=priorities.index(level)+2, value=assignees[key][level])
        total += assignees[key][level]
    ws_out.cell(row=last_row, column=last_col, value=total)
    last_row += 1

ws_out.cell(row=last_row, column=1, value="TOTALS")

for i in range(len(priorities)+1):
    total = 0
    for num in ws_out.iter_rows(min_col=i+2, max_col=i+2, min_row=2, max_row=last_row-1, values_only=True):
        if (num[0] != None):
            total += num[0]
    ws_out.cell(row=last_row, column=i+2, value=total)

wb.save("jira-reporter.xlsx")

